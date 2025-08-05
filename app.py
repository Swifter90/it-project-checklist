from flask import Flask, request, render_template, send_file, session
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd
import sqlite3
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

app = Flask(__name__)
app.secret_key = 'your-secret-key'

# Инициализация базы данных SQLite
def init_db():
    conn = sqlite3.connect('gantt_data.db')
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS tasks
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  task TEXT, phase TEXT, team TEXT,
                  start TEXT, end TEXT, progress INTEGER)''')
    conn.commit()
    conn.close()

init_db()

# Сокращённый чек-лист
checklist = [
    {"task": "Определить бизнес-цели и границы проекта", "phase": "Инициация", "team": "Проектный офис, Аналитики", "progress": 0},
    {"task": "Составить список стейкхолдеров и назначить роли", "phase": "Инициация", "team": "Проектный офис", "progress": 0},
    {"task": "Провести техническую оценку и составить ТЗ", "phase": "Инициация", "team": "IT (Продакты)", "progress": 0},
    {"task": "Разработать план проекта и бюджет", "phase": "Планирование", "team": "Проектный офис, Аналитики", "progress": 0},
    {"task": "Определить стек технологий и ресурсы", "phase": "Планирование", "team": "IT (Продакты, Разработка)", "progress": 0},
    {"task": "Разработать план обучения", "phase": "Планирование", "team": "Обучение", "progress": 0},
    {"task": "Провести кик-офф встречу", "phase": "Исполнение", "team": "Проектный офис", "progress": 0},
    {"task": "Реализовать разработку и тестирование", "phase": "Исполнение", "team": "IT (Разработка)", "progress": 0},
    {"task": "Подготовить и провести обучение", "phase": "Исполнение", "team": "Обучение", "progress": 0},
    {"task": "Участвовать в пилотном тестировании", "phase": "Исполнение", "team": "Операции", "progress": 0},
    {"task": "Отслеживать выполнение плана", "phase": "Мониторинг и контроль", "team": "Проектный офис", "progress": 0},
    {"task": "Контролировать качество", "phase": "Мониторинг и контроль", "team": "IT (Разработка)", "progress": 0},
    {"task": "Провести полное внедрение", "phase": "Завершение", "team": "Проектный офис, Операции", "progress": 0},
    {"task": "Оценить результаты и провести ретроспективу", "phase": "Завершение", "team": "Проектный офис, Аналитики", "progress": 0},
    {"task": "Закрыть проект и архивировать данные", "phase": "Завершение", "team": "Проектный офис", "progress": 0},
]

# Вехи (milestones)
milestones = [
    {"task": "Кик-офф завершен", "phase": "Исполнение", "team": "Проектный офис", "date": None},
    {"task": "Разработка завершена", "phase": "Исполнение", "team": "IT (Разработка)", "date": None},
    {"task": "Внедрение завершено", "phase": "Завершение", "team": "Проектный офис", "date": None},
]

# Обновлённые зависимости
dependencies = [
    (0, 1), (1, 3), (3, 6), (6, 7), (7, 10), (10, 12), (12, 13), (13, 14),
]

@app.route('/', methods=['GET', 'POST'])
def index():
    conn = sqlite3.connect('gantt_data.db')
    c = conn.cursor()
    c.execute('SELECT task, start, end, progress FROM tasks')
    saved_tasks = c.fetchall()
    conn.close()

    form_data = {task['task']: {'start': '', 'end': '', 'progress': 0} for task in checklist}
    for saved_task in saved_tasks:
        task_name, start, end, progress = saved_task
        form_data[task_name] = {'start': start or '', 'end': end or '', 'progress': progress or 0}

    if request.method == 'POST':
        tasks = []
        conn = sqlite3.connect('gantt_data.db')
        c = conn.cursor()
        c.execute('DELETE FROM tasks')
        for item in checklist:
            start_date = request.form.get(f'start_{item["task"]}')
            end_date = request.form.get(f'end_{item["task"]}')
            progress = request.form.get(f'progress_{item["task"]}', '0')
            progress = int(progress) if progress.isdigit() else 0
            if start_date and end_date:
                tasks.append({
                    'Task': item['task'],
                    'Phase': item['phase'],
                    'Team': item['team'],
                    'Start': start_date,
                    'Finish': end_date,
                    'Progress': progress
                })
                c.execute('INSERT INTO tasks (task, phase, team, start, end, progress) VALUES (?, ?, ?, ?, ?, ?)',
                          (item['task'], item['phase'], item['team'], start_date, end_date, progress))
        conn.commit()
        conn.close()

        # Добавляем вехи на основе введённых дат
        updated_milestones = []
        for milestone in milestones:
            if milestone['task'] == "Кик-офф завершен":
                for task in tasks:
                    if task['Task'] == "Провести кик-офф встречу":
                        milestone['date'] = task['Finish']
            elif milestone['task'] == "Разработка завершена":
                for task in tasks:
                    if task['Task'] == "Реализовать разработку и тестирование":
                        milestone['date'] = task['Finish']
            elif milestone['task'] == "Внедрение завершено":
                for task in tasks:
                    if task['Task'] == "Провести полное внедрение":
                        milestone['date'] = task['Finish']
            if milestone['date']:
                updated_milestones.append({
                    'Task': milestone['task'],
                    'Phase': milestone['phase'],
                    'Team': milestone['team'],
                    'Start': milestone['date'],
                    'Finish': milestone['date'],
                    'Progress': 100
                })

        tasks.extend(updated_milestones)

        if not tasks:
            return render_template('index.html', checklist=checklist, form_data=form_data,
                                 error="Введите даты для хотя бы одной задачи")

        session['tasks'] = tasks

        df = pd.DataFrame(tasks)
        
        fig = go.Figure()
        colors = {'Инициация': '#1f77b4', 'Планирование': '#ff7f0e', 'Исполнение': '#2ca02c',
                  'Мониторинг и контроль': '#d62728', 'Завершение': '#9467bd'}
        
        for _, row in df.iterrows():
            duration = (pd.to_datetime(row['Finish']) - pd.to_datetime(row['Start'])).days
            if duration == 0:  # Вехи
                fig.add_trace(go.Scatter(
                    x=[pd.to_datetime(row['Start'])],
                    y=[row['Task']],
                    mode='markers',
                    marker=dict(symbol='diamond', size=12, color=colors[row['Phase']], line=dict(width=2, color='black')),
                    name=row['Phase'],
                    hovertemplate=f"Веха: {row['Task']}<br>Фаза: {row['Phase']}<br>Команда: {row['Team']}<br>Дата: {row['Start']}"
                ))
            else:
                fig.add_trace(go.Bar(
                    x=[duration],
                    y=[row['Task']],
                    base=[pd.to_datetime(row['Start'])],
                    marker=dict(color=colors[row['Phase']]),
                    name=row['Phase'],
                    hovertemplate=f"Задача: {row['Task']}<br>Фаза: {row['Phase']}<br>Команда: {row['Team']}<br>Начало: {row['Start']}<br>Конец: {row['Finish']}<br>Прогресс: {row['Progress']}%"
                ))
                if row['Progress'] > 0:
                    progress_days = duration * row['Progress'] / 100
                    fig.add_trace(go.Bar(
                        x=[progress_days],
                        y=[row['Task']],
                        base=[pd.to_datetime(row['Start'])],
                        marker=dict(color='#000000', opacity=0.5),
                        showlegend=False,
                        hovertemplate=f"Прогресс: {row['Progress']}%"
                    ))

        for dep in dependencies:
            if dep[0] < len(tasks) and dep[1] < len(tasks):
                start_task = tasks[dep[0]]
                end_task = tasks[dep[1]]
                fig.add_shape(
                    type="line",
                    x0=pd.to_datetime(start_task['Finish']),
                    y0=start_task['Task'],
                    x1=pd.to_datetime(end_task['Start']),
                    y1=end_task['Task'],
                    line=dict(color="black", width=2, dash="dash"),
                    xref="x", yref="y"
                )

        fig.update_layout(
            title="Диаграмма Ганта: Запуск IT-проекта",
            xaxis=dict(title="Дата", type="date"),
            yaxis=dict(title="Задача", autorange="reversed"),
            showlegend=True,
            template="plotly_white",
            barmode='overlay',
            bargap=0.1,
            height=800
        )

        critical_tasks = set([checklist[i]['task'] for i in set(sum(dependencies, ()))])
        for trace in fig.data:
            if trace.y and trace.y[0] in critical_tasks and trace.mode != 'markers':
                trace.marker.line = dict(width=2, color="red")

        graph = fig.to_html(full_html=False)
        return render_template('index.html', checklist=checklist, form_data=form_data,
                             graph=graph, excel_available=True,
                             phases=list(set(item['phase'] for item in checklist)),
                             teams=list(set(item['team'] for item in checklist)))

    return render_template('index.html', checklist=checklist, form_data=form_data,
                         phases=list(set(item['phase'] for item in checklist)),
                         teams=list(set(item['team'] for item in checklist)))

@app.route('/download_excel', methods=['GET'])
def download_excel():
    tasks = session.get('tasks', [])
    if not tasks:
        return "Ошибка: нет данных для экспорта. Сначала постройте диаграмму.", 400

    # Создание Excel-файла
    wb = Workbook()
    ws = wb.active
    ws.title = "Диаграмма Ганта"

    # Определяем минимальную и максимальную даты
    dates = []
    for task in tasks:
        dates.append(pd.to_datetime(task['Start']))
        dates.append(pd.to_datetime(task['Finish']))
    min_date = min(dates)
    max_date = max(dates)
    date_range = pd.date_range(min_date, max_date, freq='D')

    # Заголовки
    headers = ['Задача', 'Фаза', 'Команда', 'Начало', 'Конец', 'Прогресс (%)'] + [d.strftime('%Y-%m-%d') for d in date_range]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')

    # Цвета для фаз
    phase_colors = {
        'Инициация': PatternFill(start_color='1F77B4', end_color='1F77B4', fill_type='solid'),
        'Планирование': PatternFill(start_color='FF7F0E', end_color='FF7F0E', fill_type='solid'),
        'Исполнение': PatternFill(start_color='2CA02C', end_color='2CA02C', fill_type='solid'),
        'Мониторинг и контроль': PatternFill(start_color='D62728', end_color='D62728', fill_type='solid'),
        'Завершение': PatternFill(start_color='9467BD', end_color='9467BD', fill_type='solid')
    }
    milestone_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    critical_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    progress_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

    critical_tasks = set([checklist[i]['task'] for i in set(sum(dependencies, ()))])
    
    # Заполнение данных
    for row_idx, task in enumerate(tasks, 2):
        ws.cell(row=row_idx, column=1).value = task['Task']
        ws.cell(row=row_idx, column=2).value = task['Phase']
        ws.cell(row=row_idx, column=3).value = task['Team']
        ws.cell(row=row_idx, column=4).value = task['Start']
        ws.cell(row=row_idx, column=5).value = task['Finish']
        ws.cell(row=row_idx, column=6).value = task['Progress']
        
        start_date = pd.to_datetime(task['Start'])
        end_date = pd.to_datetime(task['Finish'])
        is_milestone = start_date == end_date
        
        for col_idx, date in enumerate(date_range, 7):
            if start_date <= date <= end_date:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = milestone_fill if is_milestone else phase_colors[task['Phase']]
                if task['Task'] in critical_tasks:
                    cell.border = Border(left=Side(style='thin', color='FF0000'),
                                       right=Side(style='thin', color='FF0000'),
                                       top=Side(style='thin', color='FF0000'),
                                       bottom=Side(style='thin', color='FF0000'))
                if task['Progress'] > 0 and not is_milestone:
                    progress_days = (end_date - start_date).days * task['Progress'] / 100
                    progress_end = start_date + timedelta(days=progress_days)
                    if start_date <= date <= progress_end:
                        cell.fill = progress_fill

    # Настройка ширины столбцов
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions['A'].width = 40

    # Сохранение в буфер
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return send_file(excel_buffer, download_name="gantt_chart.xlsx", as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
